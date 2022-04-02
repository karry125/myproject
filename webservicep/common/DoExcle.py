# -*- coding:utf-8 -*-
# @Time:2019/4/13
# @Author:wangym
# @Email:123@qq.com
# @File:DoExcle

from openpyxl import load_workbook


class Cases:
    """
    测试用例类，包含以下字段
    """
    def __init__(self):
        self.id=None
        self.title=None
        self.url=None
        self.api=None
        self.params=None
        self.exceptend=None
        self.actually=None
        self.result=None
        self.sql=None


class DoExcle:

    def __init__(self,filename,sheet):
        self.filename=filename
        self.work=load_workbook(filename)
        self.sheetname=self.work[sheet]

    def getcase(self):
        """
        获取excle用例以字典格式放到list
        :return:
        """
        row =self.sheetname.max_row
        col =self.sheetname.max_column
        cases=[]
        title=[]
        # for x in range(1,col+1):
        #     coll=self.sheetname.cell(1,x).value
        #     title.append(coll)
        print(title)
        for i in range(2,row+1):
            case={ }
            print(title[0])
            for j in range(1,col+1):
                print(title[j-1])
                case[title[j-1]]=self.sheetname.cell(i,j).value
            cases.append(case)
        return cases

    def getcases(self):
        """
        测试用例以Cases类实例形式放在list中
        :return:
        """
        row = self.sheetname.max_row
        col = self.sheetname.max_column
        cases = []
        #title = []
        #标题行
        # for x in range(1, col + 1):
        #     coll = self.sheetname.cell(1, x).value
        #     title.append(coll)
        # print(title)
        for i in range(2, row + 1):
            case = Cases()
            case.id = self.sheetname.cell(i, 1).value
            case.title = self.sheetname.cell(i, 2).value
            case.url= self.sheetname.cell(i, 3).value
            case.api = self.sheetname.cell(i, 4).value
            case.params = self.sheetname.cell(i, 5).value
            case.exceptend = self.sheetname.cell(i, 6).value
            case.sql = self.sheetname.cell(i,9).value
            cases.append(case)
        return cases

    def writerow(self,row,actually,result):
        """
        测试用例的实际值跟结果写会excle，最后保存关闭
        :param row:
        :param actually:
        :param result:
        :return:
        """
        self.sheetname.cell(row,7).value = actually
        self.sheetname.cell(row, 8).value = result
        self.work.save(self.filename)
        self.work.close()


if __name__ == '__main__':
        read = DoExcle("../data/前程贷接口用例.xlsx","login")
        case = read.getcases()
        for casee in case:
            print(casee.id,casee.title)
            print(type(casee.id))




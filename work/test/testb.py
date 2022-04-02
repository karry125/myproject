# -*- coding:utf-8 -*-
# @Time:2019/3/16
# @Author:wangym
# @Email:123@qq.com
# @File:testb


#打开文件：
from openpyxl import load_workbook
excel=load_workbook('../date0314/k.xlsx')
#获取sheet：
table = excel.get_sheet_by_name('开心')   #通过表名获取
#获取行数和列数：
rows=tablemax_row   #获取行数
cols=table.max_column    #获取列数
#获取单元格值：
Data=table.cell(row=row,column=col).value  #获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value
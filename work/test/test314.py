# -*- coding:utf-8 -*-
# @Time:2019/3/14
# @Author:wangym
# @Email:123@qq.com
# @File:test314

#安排一个作业 #写一个类 类里面有2个方法 1）读数据 2）写数据

#1）读数据可以读取整个Excel里面所有的数据，
# 每一行数据都放到一个子列表里面，所有子列表数据放到一个大列表里面
# ,要求把读取到的数据返回 #2）写数据可以在Excel里面指定的单元格里面写入指定的值，不需要返回值



#温馨提示：记得关闭和保存Excel
from openpyxl import workbook
from openpyxl import load_workbook

# file = open("a.xlsx","w+",encoding="utf-8")
# print(file.readlines())


class xlsxdo():

    def readxlxs(self,workname,sheetname):
        wb = load_workbook(workname)
        shee = wb.get_sheet_by_name[sheetname]
        roww = shee.max_row
        loww = shee.max_column
        l=[]
        for i in range(1,roww+1):
            m=[]
            for j in range(1,loww+1):
                b=shee.cell(i,j).value
                m.append(b)
                print(b)
            l.append(m)
        print(l)

    def writexlsx(self,workname,sheetname,data):
        wb = workbook.Workbook()
        she= wb.create_sheet(sheetname)
        for i in range(0,len(data)):
            for j in range(0,len(data[i])):
               print(data[i][j])
               she.cell(i+1,j+1,data[i][j])
        wb.save(workname)



h = xlsxdo()
data = [["薯片","饼干"," 方便面"],[" 酸奶"," 饮料"," 开水"],[" 冰淇淋"," 雪糕"," 棒冰"]]
print(data[1][1])
h.writexlsx("b.xlsx","零售",data)
#h.readxlxs("a.xlsx","Sheet1")
wb = load_workbook("a.xlsx")
#h.writexlsx("a.xlsx")
she = wb["Sheet"]
# wbb = workbook.Workbook()
# s=wbb.create_sheet("1")
# s.cell(1,1,"12")

# wbb.save("one.xlsx")
# print(a1)
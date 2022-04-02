# -*- coding:utf-8 -*-
# @Time:2019/3/16
# @Author:wangym
# @Email:123@qq.com
# @File:xlsxhelp

# 安排一个作业 #写一个类 类里面有2个方法 1）读数据 2）写数据
# 1）读数据可以读取整个Excel里面所有的数据，每一行数据都放到一个子列表里面，
# 所有子列表数据放到一个大列表里面,要求把读取到的数据返回
#  #2）写数据可以在Excel里面指定的单元格里面写入指定的值，不需要返回值
# 温馨提示：记得关闭和保存Excel


from openpyxl import workbook
from openpyxl import load_workbook
import  time

class XlsxHelp:
    '''  表格读写方法  '''
    def __init__(self, work, sheet):
        self.work = work
        self.sheet = sheet

    def readxlxs(self):
        '''  读取表格方法
        workk 工作簿文件地址，当前目录文件名.xlsx
        sheett 哪个sheet表 '''
        l =[]
        wb = load_workbook(self.work)
        shee = wb[self.sheet]
        row = shee.max_row
        col = shee.max_column

        for i in range(2,row+1):
            m =[]
            for j in range(1,col+1):
                b = shee.cell(i,j).value
                if b:  #有空值不添加到list
                   m.append(b)
            l.append(m)
        print(l)
        return l
    def readtitle(self):
        l = []
        wb = load_workbook(self.work)
        shee = wb[self.sheet]
        row = shee.max_row
        col = shee.max_column
        rowlist=[]
        for i in range(1,col):
            m=shee.cell(1,i).value
            rowlist

    def newxlsx(self, data, title =None):
        """新建一个工作簿工作表插入标题和数据 title默认无"""
        timee = time.strftime('%Y-%m-%d_%H-%M-%S')
        excelname = self.work.split(".")[0]+ timee + '.xlsx'
        wb = workbook.Workbook()
        she = wb.create_sheet(self.sheet,0)
        if title:
            for x in range(len(title)):
                she.cell(1,x+1,title[x])
            for i in range(0,len(data)):
                for j in range(0,len(data[i])):
                    she.cell(i+2,j+1,data[i][j])
        else:
            for i in range(0,len(data)):
                for j in range(0,len(data[i])):
                    she.cell(i+1,j+1,data[i][j])
        wb.save(excelname)  # 保存文件

    def writeappend(self,data):
        """追加写"在最后一行"""
        wb = load_workbook(self.work)
        shee = wb[self.sheet]
        row = shee.max_row
        col = shee.max_column
        for i in range(row,row+len(data)):
            for j in range(0,len(data[i-row])):
                shee.cell(i+1,j+1,data[i-row][j])
        wb.save(self.work)

    def writexlsx(self,data):
        """从第一行开始写，传入list，最多两层嵌套"""
        wb = load_workbook(self.work)
        shee = wb[self.sheet]
        for i in range(0, len(data)):
            for j in range(0, len(data[i])):
                shee.cell(i + 1, j + 1, data[i][j])
        wb.save(self.work)

    def writerowxlsx(self,row,colum,data):
        """ 根据行列值写入单元格
        :row行
        :colum列
        """
        wb = load_workbook(self.work)
        shee = wb[self.sheet]
        shee.cell(row, colum, data)
        wb.save(self.work)
    def readrow(self,row,title=None):
        wb = load_workbook(self.work)
        shee = wb[self.sheet]
        row = shee.max_row
        col = shee.max_column
        rowlist=[]
        for cols in range(1,col+1):
            cell = shee.cell(row,cols).value
            rowlist.append(cell)
        return rowlist




if __name__ == '__main__':

    h = XlsxHelp("../day0321/http.xlsx","开心")
    title = ["标题1","标题2","标题3"]
    data = [["薯片g","饼干"," 方便面"],[" 酸奶"," 饮料"," 开水"],[" 冰淇淋"," 雪糕"," 棒冰"],[" 1冰淇淋"," 雪糕"," 1棒冰"]]
    # h.writexlsx(data)
    # h.writerowxlsx(5,5,"n")
    m =h.readrow(2)
    print(m)


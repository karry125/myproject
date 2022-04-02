# -*- coding:utf-8 -*-
# @Time:2019/3/14
# @Author:wangym
# @Email:123@qq.com
# @File:xlsx_help

# 安排一个作业 #写一个类 类里面有2个方法 1）读数据 2）写数据
# 1）读数据可以读取整个Excel里面所有的数据，每一行数据都放到一个子列表里面，
# 所有子列表数据放到一个大列表里面,要求把读取到的数据返回
#  #2）写数据可以在Excel里面指定的单元格里面写入指定的值，不需要返回值



#温馨提示：记得关闭和保存Excel
from openpyxl import workbook
from openpyxl import load_workbook

# file = open("a.xlsx","w+",encoding="utf-8")
# print(file.readlines())


class XlsxHelp:
    '''  表格读写方法  '''
    def __init__(self, work, sheet):
        self.work = work
        self.wb = load_workbook(work)
        self.shee = self.wb[sheet]
        self.row = self.shee.max_row
        self.col = self.shee.max_column

    def readxlxs(self, workk, sheett):
        '''  读取表格方法
        workk 工作簿文件地址，当前目录文件名.xlsx
        sheett 哪个sheet表 '''

      #get_sheet_by_name[sheett]报错TypeError: 'method' object is not subscriptable

        l=[]
        for i in range(1,self.row+1):
            m=[]
            for j in range(1,self.col+1):
                b=self.shee.cell(i,j).value
                m.append(b)
                print(b)
            l.append(m)
        print(self.shee.rows)
        print(l)


    def newxlsx(self, workk,sheet, data, title =None):
        """新建一个工作簿工作表插入标题和数据 title默认无"""
        wb = workbook.Workbook()
        she= wb.create_sheet(sheet)
        if title:
            for x in range(len(title)):
                she.cell(1,x+1,title[x])
            for i in range(0,len(data)):
                for j in range(0,len(data[i])):
                    she.cell(i+2,j+1,data[i][j])
        else:
            for i in range(0,len(data)):
                for j in range(0,len(data[i])):
                    she.cell(i+1,j+1,data[i][j])
        wb.save(workk)  # 保存文件

    def writeappend(self,data):
        """t"""
        for i in range(self.roww,self.roww+len(data)):
            for j in range(0,len(data[i-self.roww])):
                self.shee.cell(i+1,j+1,data[i-self.roww][j])
        self.wb.save(self.work)

    def writexlsx(self,data):
        """直接写入覆盖写，传入list，最多两层嵌套"""
        for i in range(0, len(data)):
            for j in range(0, len(data[i])):
                self.shee.cell(i + 1, j + 1, data[i][j])
        self.wb.save(self.work)

    def writerowxlsx(self,row,colum,data):
        """ 根据行列值写入单元格
        :row行
        :colum列
        """
        self.shee.cell(row, colum, data)
        self.wb.save(self.work)

    def creatbook(self,work,sheet):
        pass

    def openbook(self):
        wb = load_workbook(self.work)
        shee = wb[self.sheet]
        row = shee.max_row
        col = shee.max_column
        return shee, row, col, wb




h = XlsxHelp("k.xlsx","开心")
title = ["标题1","标题2","标题3"]
data = [["薯片","饼干"," 方便面"],[" 酸奶"," 饮料"," 开水"],[" 冰淇淋"," 雪糕"," 棒冰"],[" 1冰淇淋"," 雪糕"," 1棒冰"]]
#h.readxlxs("a.xlsx","Sheet1")
#h.writexlsx("k.xlsx","开心",data)
#h.readxlxs("k.xlsx","开心")
#h.newxlsx
#h.writeappend(data)
#
# w = load_workbook("k.xlsx")
# sheet = w[w.sheetnames[1]]
b =workbook.Workbook
s = b.create_sheet("mm")

b.save("o.xlsx")
n=s.max_row
print(n)
# for row in sheet.rows:
#         print(row)
#         print(sheet.rows)
#         for k in row:
#             k.value
#             print(k.value,k)
#         line = [col.value for col in row]
#
#         #print (line)
# print(type(sheet.max_row))
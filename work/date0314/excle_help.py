# -*- coding:utf-8 -*-
# @Time:2019/3/24
# @Author:wangym
# @Email:123@qq.com
# @File:excle_help



from openpyxl import workbook,load_workbook

#创建工作簿
work = workbook.Workbook()
sheetname=work.create_sheet("kk",index=0)
work.save("my.xlsx")
#读取表格，要打开工作簿
worker = load_workbook("kk.xlsx")
#找到sheetname
sheet =worker["开心"]
row =sheet.max_row  #最大行
col = sheet.max_column
for i in range(1,row+1):
    for j in range(1,col+1):
        cells=sheet.cell(i,j).value
        if cells:#空的打印None
           print(cells)


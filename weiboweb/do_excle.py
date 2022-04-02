# -*- coding:utf-8 -*-
# @Time:2019/6/11
# @Author:wangym
# @Email:123@qq.com
# @File:do_excle

from openpyxl import load_workbook
import getcookies
import httpforword
class case:
    name=None
    passs=None
    id=None
    cook=None


class do_excle():
    def __init__(self,filename,sheetname):
        self.filename = filename
        self.work = load_workbook(filename)
        self.sheetnames = self.work[sheetname]

    def read_excle(self):

        #sheetnames.cell(row,col).value = values
        maxrow=self.sheetnames.max_row
        maxcol=self.sheetnames.max_column
        ac=[]
        for i in range(1,maxrow+1):
          oneaccount = case()
          oneaccount.id=self.sheetnames.cell(i,1).value
          oneaccount.name=self.sheetnames.cell(i,2).value
          oneaccount.passs=self.sheetnames.cell(i,3).value
          oneaccount.cook=self.sheetnames.cell(i,4).value
          ac.append(oneaccount)
        self.work.close()
        print(ac)
        return ac
        # work.save(fielname)
        # work.close()


    def write_excle(self,row,col,values):
         print("uu",values)
         self.sheetnames.cell(row,col).value = values
         self.work.save(self.filename)
         self.work.close()

if __name__ == '__main__':
    fielname = 'account.xlsx'
    work = do_excle(fielname, "21")
    account = work.read_excle()
    cook={"Cookie":"SINAGLOBAL=4789045886077.128.1533964265475; YF-V5-G0=a5a6106293f9aeef5e34a2e71f04fae4; _s_tentry=login.sina.com.cn; UOR=www.228.com.cn,widget.weibo.com,login.sina.com.cn; Apache=5688356668637.786.1560668654027; ULV=1560668654044:33:4:1:5688356668637.786.1560668654027:1560350543693; wb_view_log_5317798778=1920*10801; Ugrow-G0=9ec894e3c5cc0435786b4ee8ec8a55cc; login_sid_t=e84894f2a6be4721f6762a3db9bd2899; cross_origin_proto=SSL; WBStorage=6b696629409558bc|undefined; wb_view_log=1920*10801; wb_view_log_6910869075=1920*10801; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9WhiAEVTP0f3w5FO87Rw_VV25JpX5K2hUgL.Foq4eh.4e0McSKM2dJLoIp7LxKML1KBLBKnLxKqL1hnLBoMc1K541KeNSo-N; ALF=1592205800; SSOLoginState=1560669801; SCF=AqOoUEpU3pPp0BTfcqRmHG_yVubh1Um_R5Bax4TCzB58C1FFdP4WJ9aF_ReFjGubldZnvVSCowxy6Z0owl5Ca4U.; SUB=_2A25wAZ46DeRhGeBH61sY8ynKzjuIHXVTdojyrDV8PUNbmtBeLUXwkW9NQcLKP3k9v9t8nhytSsw5VfKsCH7Ig1eE; SUHB=0oU4zVN4Yag6-d; un=ssfadndsriwaiv-svb@yahoo.com; wvr=6; YF-Page-G0=580fe01acc9791e17cca20c5fa377d00|1560669807|1560669522; wb_view_log_6909937657=1920*10801; weiboNotfication=1379357432380.357; webim_unReadCount=%7B%22time%22%3A1560669823002%2C%22dm_pub_total%22%3A0%2C%22chat_group_pc%22%3A0%2C%22allcountNum%22%3A35%2C%22msgbox%22%3A0%7D"}
    httpforword.httpfor(cook)
    # for i in account:
    #     re=i.cook
    #     httpforword.httpfor(eval(re))
    #     print(re)
# -*- coding:utf-8 -*-
# @Time:2019/3/23
# @Author:wangym
# @Email:123@qq.com
# @File:exclehelp

from openpyxl import workbook,load_workbook,worksheet

class xlsxhelp:
    def __init__(self,workbook,sheetname):
        """workname工作簿路径"""
        self.workbook =workbook
        self.sheetname = sheetname

    def getsheet(self):
        workbooks = load_workbook(self.workbook)
        sheet = workbooks[self.sheetname]
        return sheet

    def readall(self):
        '''读取全部，放列表里返回'''
        # workbooks=load_workbook(self.workbook)
        # sheet = workbooks[self.sheetname]
        # row = sheet.max_row
        # col = sheet.max_column
        sheet = self.getsheet()
        row = self.getrows()
        col = self.getcols()
        alllist=[]
        for i in range(2,row+1):
            rowlist = []
            for j in range(1,col+1):
                cell =sheet.cell(i,j).value
                if cell:
                    rowlist.append(cell)
            alllist.append(rowlist)
        print(alllist)
        return alllist

    def readcell(self,row,col):
        """读取单元格"""
        read = load_workbook(self.workbook)
        sheet = read[self.sheetname]
        cell = sheet.cell(row,col).value
        print(cell)
        return cell

    def readrowcell(self, row):
        """读取整行"""
        read = load_workbook(self.workbook)
        sheet = read[self.sheetname]
        col =sheet.max_column
        rowlist=[]
        for i in range(1,col+1):
            cell = sheet.cell(row, i).value
            rowlist.append(cell)
        print(rowlist)
        read.close()
        return rowlist

    def writecell(self,row,col,data):
        write = load_workbook(self.workbook)
        sheet = write[self.sheetname]
        sheet.cell(row,col).value=data
        write.save(self.workbook)

    def create(self):
        work = workbook.Workbook()
        work.create_sheet(self.sheetname)
        work.save(self.workbook)

    def getrows(self):
        rows = self.getsheet().max_row
        # write = load_workbook(self.workbook) 同样的方法了，写成一个方法调用
        # sheet = write[self.sheetname]
        return rows
    def getcols(self):
        cols = self.getsheet().max_column
        return cols
    def geturl(self):
        url=[]
        sheet = self.getsheet()
        for i in range(1,self.getrows()+1):
            url.append(sheet.cell(i,1).value)
        return url

if __name__ == '__main__':
    work = xlsxhelp("../data/http.xlsx","开心")
    #work.create()
    #work.writecell(1,1,"kl")
    #work.readall()
    a=work.geturl()
    print(a)